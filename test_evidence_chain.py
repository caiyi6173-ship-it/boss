import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl
from pydantic import ValidationError

import config
from boss_task_agent import LocalTableExporter
from dashboard_generator import generate_dashboard
from models import StructuredInstruction
from test_database import make_evidenced_instruction


class EvidenceChainTests(unittest.TestCase):
    def test_rejects_quote_not_found_in_source_segment(self):
        data = make_evidenced_instruction().model_dump(mode="json")
        data["tasks"][0]["evidence"][0]["source_quote"] = "原话中不存在的内容"

        with self.assertRaisesRegex(ValidationError, "引文不在 S0001 原话中"):
            StructuredInstruction(**data)

    def test_excel_export_contains_evidence_and_source_sheets(self):
        instruction = make_evidenced_instruction()
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "tasks.xlsx"
            with patch.object(config, "LOCAL_TABLE_XLSX", output):
                LocalTableExporter().export(instruction)

            workbook = openpyxl.load_workbook(output, read_only=True)
            self.assertIn("证据链", workbook.sheetnames)
            self.assertIn("来源分段", workbook.sheetnames)
            evidence_sheet = workbook["证据链"]
            source_sheet = workbook["来源分段"]
            self.assertEqual(evidence_sheet.cell(2, 4).value, "总经理")
            self.assertEqual(source_sheet.cell(2, 1).value, "S0001")
            self.assertEqual(source_sheet.cell(2, 2).value, "总经理")
            self.assertIn("先出方案", source_sheet.cell(2, 6).value)
            workbook.close()

    def test_dashboard_embeds_evidence_chain(self):
        instruction = make_evidenced_instruction()
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "dashboard.html"
            with patch.object(config, "DASHBOARD_HTML", output):
                generate_dashboard(instruction)

            html = output.read_text(encoding="utf-8")
            self.assertIn("证据链", html)
            self.assertIn("总经理", html)
            self.assertIn("先出方案", html)
            self.assertIn("00:01.200-00:08.400", html)
            self.assertIn("meeting.m4a", html)
            self.assertNotIn("{segments_json_data}", html)
            self.assertNotIn("{project_evidence_html}", html)


if __name__ == "__main__":
    unittest.main()
