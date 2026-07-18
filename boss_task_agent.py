"""
Boss Task Agent - 核心智能体
将老板语音摘要解析为结构化任务、生成 SOP、跟踪进度、生成复盘。
"""

import json
import re
import traceback
from datetime import datetime, timedelta
from pathlib import Path
from typing import Type, TypeVar

import openpyxl
from openai import OpenAI
from pydantic import BaseModel, ValidationError

import config
from database import SQLiteTaskStore
from models import (
    StructuredInstruction, ProjectOverview, Task, TaskStatus, TaskPriority,
    TaskCategory, Deliverable, SOPTemplate, SOPStep,
    ProgressReport, RetrospectiveReport, TaskProgressEntry,
    ChatIntent, ChatIntentType,
    SourceSegment,
)

T = TypeVar("T", bound=BaseModel)


# ════════════════════════════════════════════════════════════════
#  0. LLM 自我修正重试引擎 (Auto-Retry with Self-Correction)

# =================================================================
#  0.5 业务同义词词典
# =================================================================

class DomainGlossary:
    """Load domain_glossary.json (+ optional general synonym soft expand) for NLU."""

    # 通用词库 soft 扩展：只给业务标准词补短近义词，避免把整库塞进 prompt
    _GENERAL_MAX_GROUP = 12
    _GENERAL_MAX_ALIAS_LEN = 6
    _GENERAL_MAX_PER_TERM = 6
    _GENERAL_BLACKLIST = {
        "妄想", "希图", "贪图", "企图", "估计", "估量", "推算", "盘算",
        "意料", "料想", "预想", "预料", "预计", "广告", "鼓动", "声张",
        "张扬", "案牍", "壅闭", "流动", "运动", "举止", "条记", "前期",
        # 书面/生僻/跨义噪声
        "揭橥", "揭晓", "传布", "流传", "劝导", "疏导", "疏浚", "疏通",
        "质料", "原料", "片断",
    }

    def __init__(
        self,
        path: Path | str | None = None,
        general_synonym_path: Path | str | None = None,
    ):
        self.path = Path(path) if path else config.DOMAIN_GLOSSARY_JSON
        self.general_synonym_path = (
            Path(general_synonym_path)
            if general_synonym_path
            else getattr(config, "CHINESE_SYNONYM_TXT", None)
        )
        self.terms: list[dict] = []
        self.status_aliases: dict[str, list[str]] = {}
        self._alias_to_canonical: dict[str, str] = {}
        self._general_groups_loaded = 0
        self.reload()

    def reload(self) -> None:
        self.terms = []
        self.status_aliases = {}
        self._alias_to_canonical = {}
        self._general_groups_loaded = 0
        if not self.path.exists():
            print(f"   未找到业务词典: {self.path}（将跳过同义词增强）")
            return
        try:
            data = json.loads(self.path.read_text(encoding="utf-8"))
        except Exception as e:
            print(f"   业务词典读取失败: {e}")
            return

        self.terms = data.get("terms") or []
        self.status_aliases = data.get("status_aliases") or {}

        # 先装业务词，再 soft 合并通用同义词
        self._apply_general_synonyms()

        for item in self.terms:
            canonical = str(item.get("canonical") or "").strip()
            if not canonical:
                continue
            aliases = item.get("aliases") or []
            # map canonical and aliases -> canonical
            self._alias_to_canonical[canonical.lower()] = canonical
            for alias in aliases:
                a = str(alias).strip()
                if a:
                    self._alias_to_canonical[a.lower()] = canonical

        for status, aliases in self.status_aliases.items():
            for alias in aliases or []:
                a = str(alias).strip()
                if a:
                    # status aliases kept separately; not in term map
                    pass

    def _load_general_groups(self) -> list[list[str]]:
        path = self.general_synonym_path
        if not path or not Path(path).exists():
            return []
        groups: list[list[str]] = []
        try:
            for line in Path(path).read_text(encoding="utf-8").splitlines():
                parts = [p.strip() for p in line.split("\t") if p.strip()]
                if len(parts) == 1:
                    parts = [p for p in parts[0].split() if p]
                if 2 <= len(parts) <= self._GENERAL_MAX_GROUP:
                    groups.append(parts)
        except Exception as e:
            print(f"   通用同义词库读取失败: {e}")
            return []
        self._general_groups_loaded = len(groups)
        return groups

    def _apply_general_synonyms(self) -> None:
        """用通用词库给业务标准词 soft 补别名（不改磁盘 JSON）。"""
        groups = self._load_general_groups()
        if not groups or not self.terms:
            return

        index: dict[str, list[list[str]]] = {}
        for g in groups:
            for w in g:
                index.setdefault(w, []).append(g)

        for item in self.terms:
            canonical = str(item.get("canonical") or "").strip()
            if not canonical:
                continue
            aliases = [str(a).strip() for a in (item.get("aliases") or []) if str(a).strip()]
            seeds = {canonical, *aliases}
            added: list[str] = []
            seen = set(seeds)
            for seed in list(seeds):
                for group in index.get(seed, []):
                    for w in group:
                        if w in seen or w in self._GENERAL_BLACKLIST:
                            continue
                        if not (1 <= len(w) <= self._GENERAL_MAX_ALIAS_LEN):
                            continue
                        if "," in w or "，" in w or " " in w:
                            continue
                        seen.add(w)
                        added.append(w)
                        if len(added) >= self._GENERAL_MAX_PER_TERM:
                            break
                    if len(added) >= self._GENERAL_MAX_PER_TERM:
                        break
                if len(added) >= self._GENERAL_MAX_PER_TERM:
                    break
            if added:
                item["aliases"] = aliases + added

    def matched_terms(self, text: str) -> list[dict]:
        """Return glossary terms whose canonical/alias appears in text."""
        if not text or not self.terms:
            return []
        lower = text.lower()
        hits = []
        for item in self.terms:
            canonical = str(item.get("canonical") or "").strip()
            aliases = [str(a).strip() for a in (item.get("aliases") or []) if str(a).strip()]
            keys = [canonical] + aliases
            matched_alias = None
            for key in keys:
                if key and key.lower() in lower:
                    matched_alias = key
                    break
            if matched_alias:
                hits.append({
                    "canonical": canonical,
                    "matched": matched_alias,
                    "aliases": aliases,
                    "category": item.get("category", ""),
                })
        return hits

    def expand_text(self, text: str) -> str:
        """Append canonical terms for matched aliases to help retrieval."""
        hits = self.matched_terms(text)
        if not hits:
            return text
        extras = []
        for h in hits:
            extras.append(h["canonical"])
            extras.extend(h.get("aliases") or [])
        # unique preserve order
        seen = set()
        uniq = []
        for x in extras:
            k = x.lower()
            if k not in seen:
                seen.add(k)
                uniq.append(x)
        return f"{text} | 同义扩展: {' / '.join(uniq)}"

    def prompt_section(self, command: str | None = None) -> str:
        """Build a compact glossary section for LLM prompts."""
        if not self.terms and not self.status_aliases:
            return ""

        lines = ["## 业务同义词词典（匹配任务时请优先参考）"]
        if self._general_groups_loaded:
            lines.append(
                f"（已 soft 合并通用中文同义词库，共 {self._general_groups_loaded} 组近义词）"
            )
        for item in self.terms:
            canonical = item.get("canonical", "")
            aliases = " / ".join(item.get("aliases") or [])
            cat = item.get("category") or ""
            if canonical and aliases:
                prefix = f"[{cat}] " if cat else ""
                lines.append(f"- {prefix}{canonical} ≈ {aliases}")

        if self.status_aliases:
            lines.append("状态口语：")
            for status, aliases in self.status_aliases.items():
                if aliases:
                    lines.append(f"- {status} ≈ {' / '.join(aliases)}")

        if command:
            hits = self.matched_terms(command)
            if hits:
                lines.append("本次用户话命中：")
                for h in hits:
                    lines.append(
                        f"- “{h['matched']}” → 标准词“{h['canonical']}”"
                    )

        return "\n".join(lines)

    def enrich_task_brief(self, task_text: str) -> str:
        return self.expand_text(task_text)


_GLOSSARY: DomainGlossary | None = None


def get_glossary() -> DomainGlossary:
    global _GLOSSARY
    if _GLOSSARY is None:
        _GLOSSARY = DomainGlossary()
    return _GLOSSARY


# ════════════════════════════════════════════════════════════════

MAX_RETRIES = 3  # 最大重试次数


def _llm_call_with_retry(
    client: OpenAI,
    model: str,
    system_prompt: str,
    user_prompt: str,
    target_model: Type[T],
    temperature: float = 0.2,
    extra_data: dict | None = None,
) -> T:
    """
    带自我修正的 LLM 调用封装。

    流程:
      1. 调用 LLM 获取 JSON 响应
      2. 用 Pydantic 模型校验
      3. 若校验失败，将错误信息反馈给 LLM 并要求修正
      4. 最多重试 MAX_RETRIES 次

    Args:
        client: OpenAI 客户端
        model: 模型名称
        system_prompt: 系统提示词
        user_prompt: 用户提示词
        target_model: 目标 Pydantic 模型类
        temperature: 生成温度
        extra_data: 额外注入到 JSON 中的字段（如 raw_input）

    Returns:
        校验通过的 Pydantic 模型实例
    """
    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt},
    ]

    last_error = None

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            # ── Step 1: 调用 LLM ──
            response = client.chat.completions.create(
                model=model,
                messages=messages,
                temperature=temperature,
                response_format={"type": "json_object"},
            )
            content = response.choices[0].message.content

            # ── Step 2: 解析 JSON ──
            try:
                data = json.loads(content)
            except json.JSONDecodeError as e:
                raise ValueError(f"JSON 解析失败: {e}\n原始返回:\n{content[:500]}")

            # 注入额外字段
            if extra_data:
                data.update(extra_data)

            # ── Step 3: Pydantic 强类型校验 ──
            result = target_model(**data)

            if attempt > 1:
                print(f"   ✅ 第 {attempt} 次重试成功！自我修正生效。")

            return result

        except (ValidationError, ValueError, KeyError, TypeError) as e:
            last_error = e
            error_type = type(e).__name__
            error_msg = str(e)

            # 截断过长的错误信息，避免 Token 浪费
            if len(error_msg) > 800:
                error_msg = error_msg[:800] + "...(截断)"

            if attempt < MAX_RETRIES:
                print(f"   ⚠️ 第 {attempt} 次解析失败 ({error_type})，正在自我修正...")

                # ── Step 4: 构造修正指令，追加到对话历史 ──
                correction_prompt = (
                    f"你刚才返回的 JSON 数据校验失败。\n"
                    f"错误类型: {error_type}\n"
                    f"错误详情:\n{error_msg}\n\n"
                    f"请仔细检查并修正以下问题：\n"
                    f"1. 确保所有必填字段都存在且类型正确\n"
                    f"2. 枚举字段只能使用指定的值\n"
                    f"3. 数组字段不能为 null，至少为空数组 []\n"
                    f"4. 只输出修正后的完整合法 JSON，不要输出任何解释文字"
                )

                # 将 LLM 上一次的回复和修正指令加入对话上下文
                messages.append({"role": "assistant", "content": content})
                messages.append({"role": "user", "content": correction_prompt})
            else:
                print(f"   ❌ 第 {attempt} 次解析仍然失败，已达最大重试次数。")

    # 所有重试都失败，抛出最后一个错误
    raise RuntimeError(
        f"LLM 输出经过 {MAX_RETRIES} 次自我修正后仍无法通过校验。\n"
        f"最后一次错误: {last_error}"
    )


# ════════════════════════════════════════════════════════════════
#  1. Excel 读取器
# ════════════════════════════════════════════════════════════════

def _safe_label(value: str) -> str:
    return str(value).replace("[", "(").replace("]", ")").replace("\n", " ").strip()


def _format_timestamp(seconds: float) -> str:
    milliseconds = max(0, round(float(seconds) * 1000))
    minutes, remainder = divmod(milliseconds, 60_000)
    secs, millis = divmod(remainder, 1_000)
    return f"{minutes:02d}:{secs:02d}.{millis:03d}"


def format_source_segment(segment: SourceSegment) -> str:
    labels = [
        f"[{segment.id}]",
        f"[职位:{_safe_label(segment.speaker_role)}]",
    ]
    if segment.start_seconds is not None and segment.end_seconds is not None:
        labels.append(
            f"[时间:{_format_timestamp(segment.start_seconds)}-"
            f"{_format_timestamp(segment.end_seconds)}]"
        )
    labels.append(f"[来源:{_safe_label(segment.source_file)}]")
    return "".join(labels) + " " + segment.text


class ExcelReader:
    """从 Workbook.xlsx 读取老板语音转录内容"""

    def __init__(self, path: Path | str | None = None, speaker_role: str | None = None):
        self.path = Path(path) if path else config.WORKBOOK_PATH
        self.speaker_role = (speaker_role or config.DEFAULT_SPEAKER_ROLE).strip()
        self.segments: list[SourceSegment] = []

    def read(self) -> str:
        """读取全部内容并拼接为文本"""
        wb = openpyxl.load_workbook(str(self.path), read_only=True)
        ws = wb.active
        self.segments = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            for cell in row:
                if cell is not None:
                    text = str(cell).strip()
                    if text:
                        self.segments.append(
                            SourceSegment(
                                id=f"S{len(self.segments) + 1:04d}",
                                source_type="excel",
                                source_file=self.path.name,
                                speaker_role=self.speaker_role,
                                text=text,
                            )
                        )
        wb.close()
        return "\n".join(format_source_segment(item) for item in self.segments)


SUPPORTED_AUDIO_EXTENSIONS = frozenset({".mp3", ".wav", ".m4a"})
SUPPORTED_INPUT_EXTENSIONS = frozenset({".xlsx", *SUPPORTED_AUDIO_EXTENSIONS})


class AudioTranscriber:
    """使用可选的 faster-whisper 依赖在本地转写音频。"""

    def __init__(self, path: Path | str, speaker_role: str | None = None):
        self.path = Path(path)
        self.speaker_role = (speaker_role or config.DEFAULT_SPEAKER_ROLE).strip()
        self.segments: list[SourceSegment] = []

    def read(self) -> str:
        if self.path.suffix.lower() not in SUPPORTED_AUDIO_EXTENSIONS:
            supported = ", ".join(sorted(SUPPORTED_AUDIO_EXTENSIONS))
            raise ValueError(f"不支持的音频格式：{self.path.suffix}。支持：{supported}")

        try:
            from faster_whisper import WhisperModel
        except ImportError as exc:
            raise RuntimeError(
                "音频转写组件未安装。请运行：pip install -r requirements-audio.txt"
            ) from exc

        language = config.AUDIO_TRANSCRIPTION_LANGUAGE.strip().lower()
        if language in {"", "auto", "none"}:
            language = None

        print(
            "[音频转写] 正在加载本地 Whisper 模型 "
            f"{config.AUDIO_TRANSCRIPTION_MODEL}（首次使用会下载模型）..."
        )
        try:
            model = WhisperModel(
                config.AUDIO_TRANSCRIPTION_MODEL,
                device=config.AUDIO_TRANSCRIPTION_DEVICE,
                compute_type=config.AUDIO_TRANSCRIPTION_COMPUTE_TYPE,
            )
            segments, info = model.transcribe(
                str(self.path),
                language=language,
                beam_size=5,
                vad_filter=True,
                condition_on_previous_text=True,
            )
            self.segments = []
            for segment in segments:
                text = segment.text.strip()
                if not text:
                    continue
                self.segments.append(
                    SourceSegment(
                        id=f"S{len(self.segments) + 1:04d}",
                        source_type="audio",
                        source_file=self.path.name,
                        speaker_role=self.speaker_role,
                        start_seconds=float(segment.start),
                        end_seconds=float(segment.end),
                        text=text,
                    )
                )
        except Exception as exc:
            raise RuntimeError(f"音频转写失败：{exc}") from exc

        transcript = "\n".join(format_source_segment(item) for item in self.segments).strip()
        if not transcript:
            raise ValueError("音频中没有识别到有效语音，请检查文件内容或录音音量。")

        detected = getattr(info, "language", None) or language or "未知"
        print(f"[音频转写] 完成：识别语言 {detected}，共 {len(transcript)} 字")
        return transcript


class InputReader:
    """按文件扩展名读取 Excel 转写稿或本地音频。"""

    def __init__(self, path: Path | str | None = None, speaker_role: str | None = None):
        self.path = Path(path) if path else config.INPUT_PATH
        self.speaker_role = (speaker_role or config.DEFAULT_SPEAKER_ROLE).strip()
        self.segments: list[SourceSegment] = []

    def read(self) -> str:
        if not self.path.exists():
            raise FileNotFoundError(f"输入文件不存在：{self.path}")

        suffix = self.path.suffix.lower()
        if suffix == ".xlsx":
            reader = ExcelReader(self.path, self.speaker_role)
            text = reader.read()
            self.segments = reader.segments
            return text
        if suffix in SUPPORTED_AUDIO_EXTENSIONS:
            reader = AudioTranscriber(self.path, self.speaker_role)
            text = reader.read()
            self.segments = reader.segments
            return text

        supported = ", ".join(sorted(SUPPORTED_INPUT_EXTENSIONS))
        raise ValueError(f"不支持的输入格式：{suffix or '无扩展名'}。支持：{supported}")


# ════════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════════
#  2. LLM 任务解析器
# ════════════════════════════════════════════════════════════════

class TaskParser:
    """调用 OpenAI API 将自然语言转为结构化任务"""

    def __init__(self):
        self.client = OpenAI(
            api_key=config.OPENAI_API_KEY,
            base_url=config.OPENAI_BASE_URL,
        )
        self.model = config.OPENAI_MODEL

    def parse(
        self,
        raw_text: str,
        source_segments: list[SourceSegment] | None = None,
    ) -> StructuredInstruction:
        """解析语音文本 → 结构化指令（带自我修正重试）"""
        today = datetime.now().strftime("%Y-%m-%d")

        # 构造 JSON Schema 描述供 LLM 参考
        schema_hint = self._build_schema_hint()

        system_prompt = config.TASK_PARSE_SYSTEM_PROMPT.replace("{today}", today)
        user_prompt = f"""以下是带分段、职位、时间和来源标签的语音/转录内容：

---
{raw_text}
---

请将上述内容解析为结构化工作指令。今天日期是 {today}。

输出必须是合法 JSON，严格遵循以下结构：
{schema_hint}

注意：
- deadline 格式为 YYYY-MM-DD；把“今天/明天/这两天/本周/下周/月底/尽快”等口语换算成具体日期
- 一句口语可拆成多个可独立交付的任务，不要压成一个大任务
- 每个任务必须有至少一个具体交付物（文档/视频/图文/表格/确认结果等）
- 识别先后依赖：拍完再发、方案确认后执行等，写入 dependencies
- 明确负责人就填真实称呼；只说“你来/帮我”可填“执行团队”或“待分配”
- 优先级：先做/最重要/今天明天 → P0；主线任务 → P1；可往后放 → P2
- 忽略寒暄与无信息重复句
- 输入中每条信息都有 [Sxxxx] 分段 ID、职位和来源；音频还带时间范围
- overview.evidence 和每个 task.evidence 必须逐字段引用真实分段 ID
- source_quote 必须逐字复制对应分段中的原话，不得改写或拼接
- 项目目标必须有 project.objective 证据；有总体截止时必须有 project.deadline 证据
- 每个任务至少有 task 和 deliverable 证据；明确负责人、截止日期、依赖时分别补 assignee、deadline、dependency 证据
- confidence 使用 0~1 小数；明确原话可高置信，推断信息应降低置信度
- 只输出 JSON，不要输出其他内容"""

        glossary_section = get_glossary().prompt_section()
        if glossary_section:
            user_prompt = user_prompt + "\n\n" + glossary_section + "\n- 解析任务标题/描述时，把口语别名归一到标准业务词"

        source_segments = source_segments or []
        return _llm_call_with_retry(
            client=self.client,
            model=self.model,
            system_prompt=system_prompt,
            user_prompt=user_prompt,
            target_model=StructuredInstruction,
            temperature=0.2,
            extra_data={
                "raw_input": raw_text,
                "source_segments": [item.model_dump(mode="json") for item in source_segments],
            },
        )

    def _build_schema_hint(self) -> str:
        """构建供 LLM 参考的 JSON 结构提示"""
        return json.dumps({
            "overview": {
                "project_name": "项目名称",
                "background": "项目背景",
                "objective": "项目目标",
                "scope": "项目范围",
                "overall_deadline": "YYYY-MM-DD",
                "key_stakeholders": ["干系人1"],
                "evidence": [{
                    "field": "project.objective|project.deadline",
                    "source_segment_id": "S0001",
                    "source_quote": "逐字原话",
                    "confidence": 0.95
                }]
            },
            "tasks": [{
                "id": "T001",
                "title": "任务标题",
                "description": "任务详细描述",
                "category": "策划|执行|内容制作|审核|汇报|其他",
                "priority": "P0-紧急|P1-重要|P2-一般",
                "assignee": "待分配",
                "deadline": "YYYY-MM-DD",
                "deliverables": [{"name": "交付物名称", "format": "格式", "description": "描述"}],
                "status": "待开始",
                "dependencies": [],
                "notes": "",
                "evidence": [{
                    "field": "task|assignee|deadline|deliverable|dependency|priority",
                    "source_segment_id": "S0001",
                    "source_quote": "逐字原话",
                    "confidence": 0.95
                }]
            }]
        }, ensure_ascii=False, indent=2)


# ════════════════════════════════════════════════════════════════
#  2.5 对话意图理解器（自然语言 -> 结构化更新意图）
# ════════════════════════════════════════════════════════════════

class ChatIntentParser:
    """将口语化任务更新指令解析为结构化 ChatIntent。"""

    CONFIDENCE_THRESHOLD = 0.55

    def __init__(self):
        self.client = OpenAI(
            api_key=config.OPENAI_API_KEY,
            base_url=config.OPENAI_BASE_URL,
        )
        self.model = config.OPENAI_MODEL

    def parse(self, command: str, instruction: StructuredInstruction) -> ChatIntent:
        """结合当前任务清单，理解用户自然语言更新意图。"""
        today = datetime.now().strftime("%Y-%m-%d")
        tasks_brief = []
        for t in instruction.tasks:
            deliverables = "、".join(d.name for d in t.deliverables) or "无"
            search_blob = f"{t.title} {t.description} {deliverables} {t.notes}"
            glossary_hits = [
                h["canonical"] for h in get_glossary().matched_terms(search_blob)
            ]
            tasks_brief.append({
                "id": t.id,
                "title": t.title,
                "description": t.description,
                "category": t.category.value,
                "status": t.status.value,
                "progress": t.progress,
                "assignee": t.assignee,
                "deadline": t.deadline,
                "deliverables": deliverables,
                "notes": t.notes,
                "glossary_tags": glossary_hits,
            })

        schema_hint = json.dumps({
            "intent": "update_task|clarify|unknown",
            "task_id": "T001 或 null",
            "task_query": "用户指向任务的关键词",
            "status": "待开始|进行中|已完成|已延期|被阻塞|null",
            "progress": 0,
            "assignee": "负责人或 null",
            "deadline": "YYYY-MM-DD 或 null",
            "note": "备注",
            "confidence": 0.9,
            "need_clarification": False,
            "clarify_question": "",
            "candidate_task_ids": []
        }, ensure_ascii=False, indent=2)

        glossary = get_glossary()
        glossary_section = glossary.prompt_section(command)
        expanded_command = glossary.expand_text(command)
        system_prompt = config.CHAT_INTENT_SYSTEM_PROMPT.replace("{today}", today)
        user_prompt = f"""当前项目：{instruction.overview.project_name}
项目目标：{instruction.overview.objective}
今天日期：{today}

当前任务清单：
{json.dumps(tasks_brief, ensure_ascii=False, indent=2)}

{glossary_section}

用户说：
---
{command}
---

同义扩展后的用户话：
---
{expanded_command}
---

请解析用户意图，输出合法 JSON，结构如下：
{schema_hint}

要求：
- 优先用业务同义词词典把口语映射到任务（如 红薯→小红书，发抖音→抖音）
- task_id / candidate_task_ids 只能来自清单中的 id
- 若无法唯一确定任务，intent=clarify，并填写 candidate_task_ids 与 clarify_question
- 若完全无法理解，intent=unknown
- 只输出 JSON"""

        intent = _llm_call_with_retry(
            client=self.client,
            model=self.model,
            system_prompt=system_prompt,
            user_prompt=user_prompt,
            target_model=ChatIntent,
            temperature=0.1,
        )
        return self._normalize(intent, instruction, command)

    def _normalize(
        self,
        intent: ChatIntent,
        instruction: StructuredInstruction,
        command: str,
    ) -> ChatIntent:
        """校验 task_id、补全 note，并在低置信时改为 clarify。"""
        valid_ids = {t.id.upper(): t.id for t in instruction.tasks}
        task_map = {t.id: t for t in instruction.tasks}

        if intent.task_id:
            intent.task_id = valid_ids.get(intent.task_id.upper())

        cleaned_candidates = []
        for cid in intent.candidate_task_ids:
            key = str(cid).upper()
            if key in valid_ids and valid_ids[key] not in cleaned_candidates:
                cleaned_candidates.append(valid_ids[key])
        intent.candidate_task_ids = cleaned_candidates

        if not intent.note:
            intent.note = f"对话修改：{command}"
        if not intent.task_query:
            intent.task_query = command.strip()[:40]

        has_update = any([
            intent.status is not None,
            intent.progress is not None,
            intent.assignee is not None,
            intent.deadline is not None,
        ])

        if intent.intent == ChatIntentType.UPDATE_TASK:
            if (
                not intent.task_id
                or not has_update
                or intent.confidence < self.CONFIDENCE_THRESHOLD
            ):
                intent.intent = ChatIntentType.CLARIFY
                intent.need_clarification = True
                if intent.task_id and intent.task_id not in intent.candidate_task_ids:
                    intent.candidate_task_ids = [intent.task_id] + intent.candidate_task_ids
                if not intent.clarify_question:
                    intent.clarify_question = self._default_clarify_question(
                        intent, task_map, command
                    )
                intent.task_id = None

        if intent.intent == ChatIntentType.CLARIFY:
            intent.need_clarification = True
            if not intent.clarify_question:
                intent.clarify_question = self._default_clarify_question(
                    intent, task_map, command
                )

        if intent.intent == ChatIntentType.UNKNOWN:
            intent.need_clarification = False
            if not intent.clarify_question:
                intent.clarify_question = (
                    "我没听懂要改哪个任务、改成什么。"
                    "可以试试：'T003 进度 60%'、'小红书任务完成'、'拍摄卡住了'"
                )

        if intent.status == TaskStatus.COMPLETED and intent.progress is None:
            intent.progress = 100
        if intent.status == TaskStatus.PENDING and intent.progress is None:
            intent.progress = 0
        if intent.progress is not None and intent.status is None:
            intent.status = (
                TaskStatus.COMPLETED if intent.progress >= 100 else TaskStatus.IN_PROGRESS
            )

        return intent

    @staticmethod
    def _default_clarify_question(intent: ChatIntent, task_map: dict, command: str) -> str:
        if intent.candidate_task_ids:
            options = []
            for tid in intent.candidate_task_ids[:3]:
                task = task_map.get(tid)
                if task:
                    options.append(f"{tid} {task.title}")
            joined = "；".join(options)
            return f"我找到多个可能任务：{joined}。你指的是哪一个？建议带上任务编号。"
        return (
            f"我还不能确定“{command}”对应哪个任务、要怎么改。"
            "请补充任务编号或更具体的名称，例如：'T002 进度 60%'。"
        )


# ════════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════════
#  3. SOP 流程生成器
# ════════════════════════════════════════════════════════════════

class SOPGenerator:
    """从结构化任务自动生成可复用的 SOP 模板"""

    def __init__(self):
        self.client = OpenAI(
            api_key=config.OPENAI_API_KEY,
            base_url=config.OPENAI_BASE_URL,
        )
        self.model = config.OPENAI_MODEL

    def generate(self, instruction: StructuredInstruction) -> SOPTemplate:
        """根据项目任务生成 SOP 模板"""
        tasks_summary = json.dumps(
            [t.model_dump() for t in instruction.tasks],
            ensure_ascii=False, indent=2
        )

        schema_hint = json.dumps({
            "template_id": "SOP-001",
            "template_name": "模板名称",
            "category": "活动策划类",
            "applicable_scenarios": ["场景1", "场景2"],
            "steps": [{
                "step_number": 1,
                "action": "动作描述",
                "responsible": "负责角色",
                "inputs": ["输入物"],
                "outputs": ["输出物"],
                "tools": ["所需工具"],
                "quality_criteria": "质量标准",
                "estimated_hours": 2.0
            }],
            "checklist": ["检查项1"],
            "risk_points": ["风险1"],
            "created_from": instruction.overview.project_name
        }, ensure_ascii=False, indent=2)

        user_prompt = f"""以下是一个项目的结构化任务清单：

项目名称：{instruction.overview.project_name}
项目目标：{instruction.overview.objective}

任务列表：
{tasks_summary}

请基于这些任务提炼一份通用 SOP 流程模板。
- 去除项目特定细节，保留通用步骤
- 适用于同类型项目的复用

输出必须是合法 JSON，结构如下：
{schema_hint}

只输出 JSON，不要输出其他内容。"""

        return _llm_call_with_retry(
            client=self.client,
            model=self.model,
            system_prompt=config.SOP_GENERATE_SYSTEM_PROMPT,
            user_prompt=user_prompt,
            target_model=SOPTemplate,
            temperature=0.3,
            extra_data={"created_from": instruction.overview.project_name},
        )

    def save_as_markdown(self, sop: SOPTemplate) -> Path:
        """将 SOP 模板保存为 Markdown 文件"""
        filename = f"{sop.template_id}_{sop.template_name}.md"
        # 清理文件名中的特殊字符
        filename = re.sub(r'[\\/:*?"<>|]', '_', filename)
        filepath = config.SOP_DIR / filename

        lines = [
            f"# {sop.template_name}",
            f"\n> **模板编号**: {sop.template_id}  ",
            f"> **分类**: {sop.category}  ",
            f"> **版本**: {sop.version}  ",
            f"> **创建时间**: {sop.created_at}  ",
            f"> **来源项目**: {sop.created_from}",
            f"\n## 适用场景\n",
        ]
        for s in sop.applicable_scenarios:
            lines.append(f"- {s}")

        lines.append(f"\n## 操作步骤\n")
        for step in sop.steps:
            lines.append(f"### 步骤 {step.step_number}: {step.action}\n")
            lines.append(f"- **负责人**: {step.responsible}")
            if step.inputs:
                lines.append(f"- **输入**: {', '.join(step.inputs)}")
            if step.outputs:
                lines.append(f"- **输出**: {', '.join(step.outputs)}")
            if step.tools:
                lines.append(f"- **工具**: {', '.join(step.tools)}")
            if step.quality_criteria:
                lines.append(f"- **质量标准**: {step.quality_criteria}")
            if step.estimated_hours > 0:
                lines.append(f"- **预计工时**: {step.estimated_hours} 小时")
            lines.append("")

        if sop.checklist:
            lines.append("## 检查清单\n")
            for item in sop.checklist:
                lines.append(f"- [ ] {item}")
            lines.append("")

        if sop.risk_points:
            lines.append("## ⚠️ 风险提示\n")
            for risk in sop.risk_points:
                lines.append(f"- ⚠️ {risk}")
            lines.append("")

        filepath.write_text("\n".join(lines), encoding="utf-8")
        return filepath


# ════════════════════════════════════════════════════════════════
#  4. 进度跟踪器
# ════════════════════════════════════════════════════════════════

class ProgressTracker:
    """任务进度跟踪与状态更新"""

    def __init__(
        self,
        store: SQLiteTaskStore | None = None,
        migrate_legacy: bool = True,
    ):
        self.store = store or SQLiteTaskStore(
            config.DATABASE_PATH,
            legacy_tasks_path=config.TASKS_JSON,
            legacy_history_path=config.PROGRESS_HISTORY_JSON,
            migrate_legacy=migrate_legacy,
        )
        self.project_id = self.store.get_active_project_id()
        self.history = self.store.load_history(self.project_id)

    def load_tasks(self) -> StructuredInstruction | None:
        """从 SQLite 加载当前活动项目。"""
        loaded = self.store.load_active_instruction()
        if not loaded:
            self.project_id = None
            self.history = []
            return None
        instruction, self.project_id = loaded
        self.history = self.store.load_history(self.project_id)
        return instruction

    def save_tasks(
        self,
        instruction: StructuredInstruction,
        new_project: bool = False,
        event: TaskProgressEntry | None = None,
    ) -> None:
        """事务性保存任务；解析新语音时使用 new_project=True。"""
        self.project_id = self.store.save_instruction(
            instruction,
            project_id=self.project_id,
            new_project=new_project,
            event=event,
        )
        self.history = self.store.load_history(self.project_id)

    def update_task(self, task_id: str, new_status: str | None = None,
                    new_progress: int | None = None, note: str = "",
                    new_assignee: str | None = None,
                    new_deadline: str | None = None) -> Task | None:
        """更新单个任务的状态、进度、负责人或截止日期"""
        instruction = self.load_tasks()
        if not instruction:
            print("❌ 未找到任务数据，请先执行 parse 命令")
            return None

        target = None
        for task in instruction.tasks:
            if task.id == task_id:
                target = task
                break

        if not target:
            print(f"❌ 未找到任务 {task_id}")
            return None

        old_status = target.status
        old_progress = target.progress

        if new_status:
            target.status = TaskStatus(new_status)
        if new_progress is not None:
            target.progress = new_progress
        if new_assignee:
            target.assignee = new_assignee
        if new_deadline:
            target.deadline = new_deadline

        target.updated_at = datetime.now().strftime("%Y-%m-%d %H:%M")

        # 记录进度历史
        entry = TaskProgressEntry(
            task_id=task_id,
            old_status=old_status,
            new_status=target.status,
            old_progress=old_progress,
            new_progress=target.progress,
            note=note,
        )
        self.save_tasks(instruction, event=entry)

        print(f"[任务更新] {task_id}: {old_status.value} → {target.status.value}, "
              f"进度 {old_progress}% → {target.progress}%")
        return target

    def get_progress_report(self) -> ProgressReport | None:
        """生成进度报告"""
        instruction = self.load_tasks()
        if not instruction:
            return None

        tasks = instruction.tasks
        total = len(tasks)
        completed = sum(1 for t in tasks if t.status == TaskStatus.COMPLETED)
        in_progress = sum(1 for t in tasks if t.status == TaskStatus.IN_PROGRESS)
        pending = sum(1 for t in tasks if t.status == TaskStatus.PENDING)
        delayed = sum(1 for t in tasks if t.status == TaskStatus.DELAYED)

        overall = sum(t.progress for t in tasks) / total if total > 0 else 0

        details = []
        for t in tasks:
            details.append({
                "id": t.id,
                "title": t.title,
                "status": t.status.value,
                "progress": t.progress,
                "priority": t.priority.value,
                "deadline": t.deadline or "未设定",
                "assignee": t.assignee,
            })

        # 找出阻塞项和亮点
        blockers = [f"{t.id} {t.title}: 已延期" for t in tasks if t.status == TaskStatus.DELAYED]
        highlights = [f"{t.id} {t.title}: 已完成" for t in tasks if t.status == TaskStatus.COMPLETED]

        return ProgressReport(
            project_name=instruction.overview.project_name,
            total_tasks=total,
            completed=completed,
            in_progress=in_progress,
            pending=pending,
            delayed=delayed,
            overall_progress=round(overall, 1),
            task_details=details,
            highlights=highlights,
            blockers=blockers,
        )


# ════════════════════════════════════════════════════════════════
#  5. 本地表格生成器（替代飞书多维表格）
# ════════════════════════════════════════════════════════════════

class LocalTableExporter:
    """将任务导出为本地 Excel 表格"""

    def export(self, instruction: StructuredInstruction) -> Path:
        """导出任务到本地 Excel"""
        wb = openpyxl.Workbook()

        # ── Sheet 1: 任务清单 ──
        ws = wb.active
        ws.title = "任务清单"

        headers = ["任务ID", "标题", "描述", "分类", "优先级", "负责人",
                    "截止日期", "状态", "进度%", "交付物", "依赖任务", "证据数", "备注",
                    "创建时间", "更新时间"]
        ws.append(headers)

        # 设置表头样式
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 写入数据
        status_colors = {
            "待开始": "FFF2CC",
            "进行中": "D6E4F0",
            "已完成": "C6EFCE",
            "已延期": "FFC7CE",
            "被阻塞": "F4B084",
        }

        for task in instruction.tasks:
            deliverables_str = "; ".join([d.name for d in task.deliverables])
            deps_str = ", ".join(task.dependencies) if task.dependencies else ""

            row_data = [
                task.id, task.title, task.description, task.category.value,
                task.priority.value, task.assignee, task.deadline or "",
                task.status.value, task.progress, deliverables_str, deps_str,
                len(task.evidence), task.notes, task.created_at, task.updated_at
            ]
            ws.append(row_data)

            # 状态列着色
            row_num = ws.max_row
            status_cell = ws.cell(row=row_num, column=8)
            color = status_colors.get(task.status.value, "FFFFFF")
            status_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            # 设置边框
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_idx).border = thin_border

        # 调整列宽
        col_widths = [8, 20, 30, 10, 10, 10, 12, 10, 8, 25, 12, 8, 20, 16, 16]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # ── Sheet 2: 项目概览 ──
        ws2 = wb.create_sheet("项目概览")
        overview_data = [
            ["项目名称", instruction.overview.project_name],
            ["项目目标", instruction.overview.objective],
            ["项目背景", instruction.overview.background],
            ["项目范围", instruction.overview.scope],
            ["总体截止", instruction.overview.overall_deadline or ""],
            ["干系人", ", ".join(instruction.overview.key_stakeholders)],
            ["解析时间", instruction.parsed_at],
        ]
        for row in overview_data:
            ws2.append(row)

        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 50

        for row_num in range(1, len(overview_data) + 1):
            ws2.cell(row=row_num, column=1).font = Font(bold=True)
            ws2.cell(row=row_num, column=1).fill = PatternFill(
                start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
            )

        # ── Sheet 3: 字段级证据链 ──
        evidence_ws = wb.create_sheet("证据链")
        evidence_headers = [
            "对象", "字段", "分段ID", "职位", "时间", "来源文件", "原话", "置信度"
        ]
        evidence_ws.append(evidence_headers)
        for col_idx in range(1, len(evidence_headers) + 1):
            cell = evidence_ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        segment_map = {item.id: item for item in instruction.source_segments}

        def append_evidence(owner: str, items):
            for item in items:
                segment = segment_map.get(item.source_segment_id)
                time_range = ""
                if segment and segment.start_seconds is not None and segment.end_seconds is not None:
                    time_range = (
                        f"{_format_timestamp(segment.start_seconds)}-"
                        f"{_format_timestamp(segment.end_seconds)}"
                    )
                evidence_ws.append([
                    owner,
                    item.field.value,
                    item.source_segment_id,
                    segment.speaker_role if segment else "",
                    time_range,
                    segment.source_file if segment else "",
                    item.source_quote,
                    item.confidence,
                ])

        append_evidence("项目概览", instruction.overview.evidence)
        for task in instruction.tasks:
            append_evidence(f"{task.id} {task.title}", task.evidence)

        for row in evidence_ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        evidence_widths = [22, 20, 12, 14, 22, 24, 55, 10]
        for i, width in enumerate(evidence_widths, 1):
            evidence_ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # ── Sheet 4: 原始来源分段 ──
        source_ws = wb.create_sheet("来源分段")
        source_headers = ["分段ID", "职位", "时间", "来源类型", "来源文件", "原话"]
        source_ws.append(source_headers)
        for col_idx in range(1, len(source_headers) + 1):
            cell = source_ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        for segment in instruction.source_segments:
            time_range = ""
            if segment.start_seconds is not None and segment.end_seconds is not None:
                time_range = (
                    f"{_format_timestamp(segment.start_seconds)}-"
                    f"{_format_timestamp(segment.end_seconds)}"
                )
            source_ws.append([
                segment.id,
                segment.speaker_role,
                time_range,
                segment.source_type,
                segment.source_file,
                segment.text,
            ])
        source_widths = [12, 14, 22, 12, 24, 70]
        for i, width in enumerate(source_widths, 1):
            source_ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        for row in source_ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        output_path = config.LOCAL_TABLE_XLSX
        wb.save(str(output_path))
        return output_path


# ════════════════════════════════════════════════════════════════
#  6. 复盘报告生成器
# ════════════════════════════════════════════════════════════════

class RetrospectiveGenerator:
    """生成复盘报告：LLM → 结构化 RetrospectiveReport → Markdown 渲染"""

    def __init__(self):
        self.client = OpenAI(
            api_key=config.OPENAI_API_KEY,
            base_url=config.OPENAI_BASE_URL,
        )
        self.model = config.OPENAI_MODEL

    def generate(self, instruction: StructuredInstruction,
                 history: list[TaskProgressEntry]) -> tuple[RetrospectiveReport, str]:
        """生成结构化复盘报告并渲染为 Markdown。

        Returns:
            (report, markdown_text)
        """
        tasks_data = json.dumps(
            [t.model_dump() for t in instruction.tasks],
            ensure_ascii=False, indent=2
        )
        history_data = json.dumps(
            [h.model_dump() for h in history],
            ensure_ascii=False, indent=2
        )

        schema_hint = json.dumps({
            "project_name": instruction.overview.project_name,
            "report_date": "YYYY-MM-DD",
            "summary": "项目总体总结，2-4 句话",
            "what_went_well": ["做得好的点1", "做得好的点2"],
            "what_needs_improvement": ["需改进点1", "需改进点2"],
            "action_items": ["具体行动项1", "具体行动项2"],
            "lessons_learned": ["经验教训1", "经验教训2"],
            "timeline_analysis": {
                "on_time_tasks": 0,
                "delayed_tasks": 0,
                "bottleneck": "关键瓶颈描述",
                "notes": "时间线其他观察"
            },
            "progress_history": []
        }, ensure_ascii=False, indent=2)

        user_prompt = f"""请为以下项目生成结构化复盘报告：

## 项目信息
- 项目名称：{instruction.overview.project_name}
- 项目目标：{instruction.overview.objective}
- 总体截止：{instruction.overview.overall_deadline or '未设定'}

## 任务执行情况
{tasks_data}

## 进度变更历史
{history_data}

输出必须是合法 JSON，严格遵循以下结构：
{schema_hint}

要求：
- 每个列表字段至少 2 条，具体且可操作，避免空话
- timeline_analysis 中的数字必须从任务数据中真实统计
- progress_history 字段固定输出空数组 []（由系统自动注入完整记录）
- 只输出 JSON，不要输出解释或 Markdown"""

        history_payload = [h.model_dump(mode="json") for h in history]

        report = _llm_call_with_retry(
            client=self.client,
            model=self.model,
            system_prompt=config.RETROSPECTIVE_SYSTEM_PROMPT,
            user_prompt=user_prompt,
            target_model=RetrospectiveReport,
            temperature=0.4,
            extra_data={
                "project_name": instruction.overview.project_name,
                "progress_history": history_payload,
            },
        )

        markdown_text = self._render_markdown(report)
        config.RETROSPECTIVE_MD.write_text(markdown_text, encoding="utf-8")
        return report, markdown_text

    @staticmethod
    def _render_markdown(report: RetrospectiveReport) -> str:
        """将结构化报告渲染为可读的 Markdown。"""
        lines = [
            f"# {report.project_name} 复盘报告",
            "",
            f"> 报告日期：{report.report_date}",
            "",
            "## 项目总结",
            "",
            report.summary or "_（无）_",
            "",
            "## 做得好的方面",
            "",
        ]
        lines += [f"- {item}" for item in report.what_went_well] or ["_（无）_"]
        lines += ["", "## 需要改进的方面", ""]
        lines += [f"- {item}" for item in report.what_needs_improvement] or ["_（无）_"]
        lines += ["", "## 经验教训", ""]
        lines += [f"- {item}" for item in report.lessons_learned] or ["_（无）_"]
        lines += ["", "## 后续行动项", ""]
        lines += [f"- [ ] {item}" for item in report.action_items] or ["_（无）_"]

        lines += ["", "## 时间线分析", ""]
        if report.timeline_analysis:
            for k, v in report.timeline_analysis.items():
                lines.append(f"- **{k}**: {v}")
        else:
            lines.append("_（无）_")

        if report.progress_history:
            lines += ["", "## 进度变更历史", ""]
            lines.append("| 时间 | 任务 | 状态变化 | 进度变化 | 备注 |")
            lines.append("|---|---|---|---|---|")
            for h in report.progress_history:
                status_change = f"{h.old_status.value} → {h.new_status.value}"
                progress_change = f"{h.old_progress}% → {h.new_progress}%"
                lines.append(
                    f"| {h.timestamp} | {h.task_id} | {status_change} | {progress_change} | {h.note} |"
                )

        return "\n".join(lines) + "\n"
